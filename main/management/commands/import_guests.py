import os
import logging
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from django.db import transaction
from main.models import Guest, GuestList
from django.conf import settings

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class Command(BaseCommand):
    help = 'Imports guests from an Excel file.'

    def add_arguments(self, parser):
        parser.add_argument('filename', type=str, help='The full path to the Excel file.')

    def handle(self, *args, **options):
        filename = options['filename']
        self.stdout.write(f"Attempting to load workbook from: {filename}")

        row_number = 1

        try:
            wb = load_workbook(filename)
            ws = wb.active

            guest_list, created = GuestList.objects.get_or_create(
                name='Main Guest List',
                defaults={'description': 'Auto-created main guest list'}
            )

            # First pass to find the maximum ID in the file
            max_id = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) > 9 and row[9] is not None:
                    try:
                        current_id = int(row[9])
                        max_id = max(max_id, current_id)
                    except ValueError:
                        continue

            next_id = max_id + 1

            # Delete all existing guests
            Guest.objects.all().delete()
            logging.info("All existing guests have been deleted.")

            with transaction.atomic():
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_number += 1
                    attending = row[4] if len(row) > 4 and isinstance(row[4], bool) else None if len(row) > 4 and not row[4] else str(row[4]).strip().lower() in ['yes', 'true', '1'] if len(row) > 4 and row[4] else False
                    email_sent_value = str(row[8]).strip().lower() if len(row) > 8 and row[8] is not None else 'false'
                    email_sent = email_sent_value in ['yes', 'true', '1']
                    message = str(row[7]).strip() if row[7] else None
                    name = row[0].strip() if row[0] else ""

                    # Determine the identification number
                    if len(row) > 9 and row[9] is not None:
                        try:
                            identification = int(row[9])
                        except ValueError:
                            identification = next_id
                            next_id += 1
                    else:
                        identification = next_id
                        next_id += 1

                    link = f"https://wedding-sjyr.onrender.com/invitation/{name}/{identification}"

                    logging.info(f"Processing row {row_number}: {row}")
                    logging.info(f"Parsed values - Attending: {attending}, Email Sent: {email_sent}, ID: {identification}")

                    guest = Guest.objects.create(
                        guest_list=guest_list,
                        name=name,
                        email=row[1].strip(),
                        phone=str(row[2]).strip() if row[2] else "",
                        country=row[3].strip() if row[3] else "",
                        attending=attending,
                        number_of_guests_invited=int(row[5]) if row[5] else 0,
                        number_of_guests_attending=int(row[6]) if row[6] else 0,
                        message=message,
                        email_sent=email_sent,
                        identification=identification,
                        link=link
                    )

            self.stdout.write(self.style.SUCCESS('Guests imported successfully.'))
        except Exception as e:
            logging.error(f"Failed to load or process the workbook at row {row_number}: {e}")
            import traceback
            traceback.print_exc()
